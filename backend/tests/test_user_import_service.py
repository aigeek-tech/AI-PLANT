import io
import unittest

from openpyxl import Workbook, load_workbook

from app.user_imports import (
    IMPORT_SHEET,
    MASKED_PASSWORD_DISPLAY,
    ROLE_REFERENCE_SHEET,
    build_user_export_workbook,
    build_user_import_template,
    validate_user_import_workbook,
)


def make_system_roles() -> list[dict]:
    return [
        {
            "id": "role-standard-admin",
            "code": "standard_admin",
            "name": "Standard Admin",
            "scope_kind": "system",
            "is_builtin": True,
            "status": "active",
            "permissions": ["standard.read", "standard.write"],
        },
        {
            "id": "role-project-creator",
            "code": "project_creator",
            "name": "Project Creator",
            "scope_kind": "system",
            "is_builtin": True,
            "status": "active",
            "permissions": ["project.create"],
        },
    ]


def make_existing_users() -> list[dict]:
    return [
        {
            "id": "user-1",
            "username": "alice",
            "email": "alice@example.test",
            "display_name": "Alice",
            "status": "active",
            "last_login_at": None,
            "created_at": "2026-04-23T00:00:00Z",
            "updated_at": "2026-04-23T00:00:00Z",
            "role_codes": ["project_creator"],
            "role_names": ["Project Creator"],
        },
        {
            "id": "user-2",
            "username": "bob",
            "email": "bob@example.test",
            "display_name": "Bob",
            "status": "active",
            "last_login_at": None,
            "created_at": "2026-04-23T00:00:00Z",
            "updated_at": "2026-04-23T00:00:00Z",
            "role_codes": ["standard_admin"],
            "role_names": ["Standard Admin"],
        },
    ]


def make_workbook(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = IMPORT_SHEET
    sheet.append(["username", "display_name", "email", "status", "password", "system_role_codes"])
    for row in rows:
        sheet.append(row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class UserImportTemplateTest(unittest.TestCase):
    def test_builds_template_and_export_with_role_reference(self):
        template_content = build_user_import_template(make_system_roles())
        template_workbook = load_workbook(io.BytesIO(template_content))

        self.assertIn("说明", template_workbook.sheetnames)
        self.assertIn(IMPORT_SHEET, template_workbook.sheetnames)
        self.assertIn(ROLE_REFERENCE_SHEET, template_workbook.sheetnames)
        template_headers = [cell.value for cell in template_workbook[IMPORT_SHEET][1]]
        self.assertEqual(
            template_headers,
            ["username", "display_name", "email", "status", "password", "system_role_codes"],
        )

        export_content = build_user_export_workbook(make_existing_users(), make_system_roles())
        export_workbook = load_workbook(io.BytesIO(export_content))
        export_sheet = export_workbook["用户导出"]
        export_headers = [cell.value for cell in export_sheet[1]]
        self.assertEqual(
            export_headers,
            [
                "username",
                "display_name",
                "email",
                "status",
                "system_role_codes",
                "system_role_names",
                "last_login_at",
                "created_at",
                "updated_at",
                "password",
            ],
        )
        self.assertEqual(export_sheet["A2"].value, "alice")
        self.assertEqual(export_sheet["E2"].value, "project_creator")
        self.assertEqual(export_sheet["F2"].value, "Project Creator")
        self.assertIn(export_sheet["J2"].value, (None, ""))


class UserImportValidationTest(unittest.TestCase):
    def test_marks_invalid_rows_as_errors(self):
        workbook_bytes = make_workbook(
            [
                ["newuser", "", "newuser@example.test", "active", "", "standard_admin"],
                ["newuser", "New User 2", "bad-email", "paused", "pass-12345", "unknown_role"],
            ]
        )

        result = validate_user_import_workbook(
            workbook_bytes,
            existing_users=make_existing_users(),
            system_roles=make_system_roles(),
            allow_role_management=True,
        )

        self.assertEqual(result["summary"]["total_rows"], 2)
        self.assertEqual(result["summary"]["error_rows"], 2)
        first_row = result["rows"][0]
        second_row = result["rows"][1]

        self.assertEqual(first_row["action"], "create")
        self.assertTrue(any(issue["field"] == "display_name" for issue in first_row["issues"]))
        self.assertTrue(any(issue["field"] == "password" for issue in first_row["issues"]))
        self.assertTrue(any(issue["code"] == "duplicate_username_in_file" for issue in second_row["issues"]))
        self.assertTrue(any(issue["field"] == "email" for issue in second_row["issues"]))
        self.assertTrue(any(issue["field"] == "status" for issue in second_row["issues"]))
        self.assertTrue(any(issue["code"] == "unknown_role_code" for issue in second_row["issues"]))

    def test_marks_high_impact_updates_as_warning_and_masks_password(self):
        workbook_bytes = make_workbook(
            [
                ["alice", "Alice Updated", "alice@example.test", "disabled", "new-password-123", "standard_admin"],
            ]
        )

        result = validate_user_import_workbook(
            workbook_bytes,
            existing_users=make_existing_users(),
            system_roles=make_system_roles(),
            allow_role_management=True,
        )

        row = result["rows"][0]
        self.assertEqual(row["action"], "update")
        self.assertEqual(row["status"], "warning")
        self.assertTrue(row["normalized_values"]["password_supplied"])
        self.assertEqual(row["values"]["password"], MASKED_PASSWORD_DISPLAY)
        self.assertEqual(row["existing_user"]["id"], "user-1")
        self.assertTrue(any(issue["code"] == "password_reset" for issue in row["issues"]))
        self.assertTrue(any(issue["code"] == "will_disable_user" for issue in row["issues"]))
        self.assertTrue(any(issue["code"] == "system_roles_changed" for issue in row["issues"]))

    def test_blocks_role_changes_without_role_permission(self):
        workbook_bytes = make_workbook(
            [
                ["newuser", "New User", "newuser@example.test", "active", "pass-12345", "standard_admin"],
            ]
        )

        result = validate_user_import_workbook(
            workbook_bytes,
            existing_users=make_existing_users(),
            system_roles=make_system_roles(),
            allow_role_management=False,
        )

        row = result["rows"][0]
        self.assertEqual(row["status"], "error")
        self.assertTrue(any(issue["code"] == "role_management_forbidden" for issue in row["issues"]))


if __name__ == "__main__":
    unittest.main()
