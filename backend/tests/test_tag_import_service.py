import io
import unittest

from openpyxl import Workbook, load_workbook

from app.tag_imports import build_tag_import_template, validate_tag_import_workbook


def make_standard_detail() -> dict:
    return {
        "id": "standard-1",
        "code": "DEC",
        "name": "DEC Standard",
        "common_attributes": [
            {
                "id": "attr-common-pressure",
                "standard_id": "standard-1",
                "class_id": None,
                "code": "pressure",
                "name": "压力",
                "value_type": "number",
                "is_required": True,
                "enum_options": [],
                "description": None,
            }
        ],
        "classes": [
            {
                "id": "class-pump",
                "code": "PUMP",
                "name": "泵",
                "attributes": [
                    {
                        "id": "attr-service",
                        "class_id": "class-pump",
                        "standard_id": None,
                        "code": "service",
                        "name": "服务",
                        "value_type": "enum",
                        "is_required": True,
                        "enum_options": ["FEED", "UTILITY"],
                        "description": None,
                    },
                    {
                        "id": "attr-speed",
                        "class_id": "class-pump",
                        "standard_id": None,
                        "code": "speed",
                        "name": "转速",
                        "value_type": "integer",
                        "is_required": False,
                        "enum_options": [],
                        "description": None,
                    },
                ],
            },
            {
                "id": "class-valve",
                "code": "VALVE",
                "name": "控制/调节阀",
                "attributes": [
                    {
                        "id": "attr-fail-position",
                        "class_id": "class-valve",
                        "standard_id": None,
                        "code": "fail_position",
                        "name": "故障位",
                        "value_type": "enum",
                        "is_required": True,
                        "enum_options": ["FO", "FC"],
                        "description": None,
                    },
                ],
            },
        ],
    }


def make_pbs_nodes() -> list[dict]:
    return [
        {"id": "pbs-1", "project_id": "project-1", "code": "UNIT-100", "name": "装置100"},
        {"id": "pbs-2", "project_id": "project-1", "code": "UNIT-200", "name": "装置200"},
    ]


def make_workbook_by_class(pump_rows: list[list[object]], valve_rows: list[list[object]] | None = None) -> bytes:
    workbook = Workbook()
    meta_sheet = workbook.active
    meta_sheet.title = "_META"
    meta_sheet.append(["sheet_name", "class_code", "class_name"])
    meta_sheet.append(["PUMP_泵", "PUMP", "泵"])
    meta_sheet.append(["VALVE_控制_调节阀", "VALVE", "控制/调节阀"])

    pump_sheet = workbook.create_sheet("PUMP_泵")
    pump_sheet.append(["tag_no", "name", "pbs_code", "压力 [pressure]", "服务 [service]", "转速 [speed]", "Tag Number [tag_no]"])
    for row in pump_rows:
        pump_sheet.append(row)

    valve_sheet = workbook.create_sheet("VALVE_控制_调节阀")
    valve_sheet.append(["tag_no", "name", "pbs_code", "压力 [pressure]", "故障位 [fail_position]"])
    for row in valve_rows or []:
        valve_sheet.append(row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class TagImportTemplateTest(unittest.TestCase):
    def test_builds_template_with_reference_sheets(self):
        content = build_tag_import_template(
            project={"code": "PRJ-001", "name": "示例项目"},
            standard_detail=make_standard_detail(),
            pbs_nodes=make_pbs_nodes(),
        )

        workbook = load_workbook(io.BytesIO(content))

        self.assertIn("说明", workbook.sheetnames)
        self.assertIn("索引", workbook.sheetnames)
        self.assertIn("PBS参考", workbook.sheetnames)
        self.assertIn("Class参考", workbook.sheetnames)
        self.assertIn("_META", workbook.sheetnames)
        self.assertEqual(workbook["_META"].sheet_state, "hidden")

        meta_rows = list(workbook["_META"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(meta_rows[0][:3], ("PUMP_泵", "PUMP", "泵"))
        self.assertEqual(meta_rows[1][:3], ("VALVE_控制_调节阀", "VALVE", "控制/调节阀"))

        data_sheet = workbook["PUMP_泵"]
        headers = [cell.value for cell in data_sheet[1]]
        self.assertEqual(
            headers,
            ["tag_no", "name", "pbs_code", "压力 [pressure]", "服务 [service]", "转速 [speed]"],
        )
        self.assertNotIn("Tag Number [tag_no]", headers)
        data_validations = list(data_sheet.data_validations.dataValidation)
        self.assertTrue(any(validation.sqref == "C2:C5001" for validation in data_validations))
        self.assertTrue(any("E2:E5001" in str(validation.sqref) for validation in data_validations))


class TagImportValidationTest(unittest.TestCase):
    def test_marks_excel_duplicates_and_missing_required_fields_as_errors(self):
        workbook_bytes = make_workbook_by_class(
            pump_rows=[
                ["P-1001", "进料泵", "UNIT-100", None, "FEED", 2900, None],
                ["P-1001", "备用泵", "UNIT-999", 0.4, "BAD", "fast", None],
            ],
        )

        result = validate_tag_import_workbook(
            workbook_bytes,
            standard_detail=make_standard_detail(),
            pbs_nodes=make_pbs_nodes(),
            existing_tags=[],
        )

        self.assertEqual(result["summary"]["total_rows"], 2)
        self.assertEqual(result["summary"]["error_rows"], 2)
        first_row = result["rows"][0]
        second_row = result["rows"][1]
        self.assertEqual(first_row["status"], "error")
        self.assertEqual(second_row["status"], "error")
        self.assertTrue(any(issue["code"] == "duplicate_in_file" for issue in first_row["issues"]))
        self.assertTrue(any(issue["code"] == "required" for issue in first_row["issues"]))
        self.assertTrue(any(issue["code"] == "pbs_not_found" for issue in second_row["issues"]))
        self.assertTrue(any(issue["code"] == "enum_invalid" for issue in second_row["issues"]))
        self.assertTrue(any(issue["code"] == "type_invalid" for issue in second_row["issues"]))

    def test_marks_existing_tag_as_conflict_and_missing_class_as_warning(self):
        workbook_bytes = make_workbook_by_class(
            pump_rows=[["P-2001", "循环泵", "UNIT-100", 0.8, "FEED", 2900, None]],
            valve_rows=[["FV-3001", "控制阀", "UNIT-200", 0.2, "FO"]],
        )

        result = validate_tag_import_workbook(
            workbook_bytes,
            standard_detail=make_standard_detail(),
            pbs_nodes=make_pbs_nodes(),
            existing_tags=[
                {
                    "id": "tag-existing",
                    "project_id": "project-1",
                    "tag_no": "P-2001",
                    "name": "旧循环泵",
                    "pbs_node_id": "pbs-1",
                    "class_id": "class-pump",
                    "attribute_values": {"pressure": 0.5},
                    "status": "active",
                }
            ],
        )

        self.assertEqual(result["summary"]["conflict_rows"], 1)
        self.assertEqual(result["summary"]["ready_rows"], 1)
        conflict_row = result["rows"][0]
        ready_row = result["rows"][1]
        self.assertEqual(conflict_row["status"], "conflict")
        self.assertEqual(conflict_row["existing_tag"]["id"], "tag-existing")
        self.assertEqual(ready_row["status"], "ready")
        self.assertEqual(ready_row["normalized_values"]["class_code"], "VALVE")
        self.assertEqual(ready_row["normalized_values"]["class_id"], "class-valve")

    def test_fixed_tag_no_column_wins_when_attribute_header_reuses_tag_no_code(self):
        workbook_bytes = make_workbook_by_class(
            pump_rows=[["P-4001", "冲突表头测试", "UNIT-100", 1.0, "FEED", 2900, None]],
        )

        result = validate_tag_import_workbook(
            workbook_bytes,
            standard_detail=make_standard_detail(),
            pbs_nodes=make_pbs_nodes(),
            existing_tags=[],
        )

        pump_row = result["rows"][0]
        self.assertEqual(pump_row["values"]["tag_no"], "P-4001")
        self.assertEqual(pump_row["normalized_values"]["tag_no"], "P-4001")
        self.assertNotIn("required", [issue["code"] for issue in pump_row["issues"]])


if __name__ == "__main__":
    unittest.main()
