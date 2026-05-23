from openpyxl import load_workbook

from app.i18n import normalize_locale, translate
from app.user_imports import _parse_user_import_workbook, build_user_import_template


def test_normalize_locale_supports_chinese_and_english_headers():
    assert normalize_locale("en-US,en;q=0.9") == "en-US"
    assert normalize_locale("zh-CN,zh;q=0.9") == "zh-CN"
    assert normalize_locale("fr-FR") == "zh-CN"
    assert normalize_locale(None) == "zh-CN"


def test_translate_falls_back_to_default_locale():
    assert translate("authInvalidCredentials", "en-US") == "Invalid username or password."
    assert translate("authInvalidCredentials", "fr-FR") == "账号或密码无效。"
    assert translate("unknownCode", "en-US") == "unknownCode"


def test_user_import_template_uses_requested_locale(tmp_path):
    workbook_path = tmp_path / "template.xlsx"
    workbook_path.write_bytes(build_user_import_template(system_roles=[], locale="en-US"))

    workbook = load_workbook(workbook_path)

    assert "User Import" in workbook.sheetnames
    assert "Role Reference" in workbook.sheetnames
    import_sheet = workbook["User Import"]
    assert [cell.value for cell in import_sheet[1]] == [
        "Username",
        "Display name",
        "Email",
        "Status",
        "Password",
        "System role codes",
    ]


def test_user_import_parser_accepts_localized_template_headers():
    rows = _parse_user_import_workbook(build_user_import_template(system_roles=[], locale="en-US"))

    assert rows[0]["username"] == "new.user"
    assert rows[0]["display_name"] == "New User"
    assert rows[0]["system_role_codes"] == "project_creator"
