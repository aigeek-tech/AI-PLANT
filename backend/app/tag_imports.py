import io
import json
import math
from collections import Counter
from datetime import date, datetime
from json import JSONDecodeError
from typing import Any, Literal
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import quote_sheetname
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from psycopg.types.json import Json

from .db import fetch_one, get_connection
from .repository import (
    get_pbs_nodes,
    get_project_detail,
    get_project_tags,
    get_standard_detail,
)


INSTRUCTION_SHEET = "说明"
INDEX_SHEET = "索引"
PBS_REFERENCE_SHEET = "PBS参考"
CLASS_REFERENCE_SHEET = "Class参考"
META_SHEET = "_META"
DEFAULT_PAGE_SIZE = 50
MAX_PAGE_SIZE = 200
MAX_UPLOAD_SIZE = 10 * 1024 * 1024
TEMPLATE_MAX_ROWS = 5001
PBS_CODES_RANGE_NAME = "pbs_codes"

REQUIRED_HEADERS = ("tag_no", "name", "pbs_code", "class_code")
CLASS_SHEET_REQUIRED_HEADERS = ("tag_no", "name", "pbs_code")
CONFLICT_ACTIONS = {"update", "skip"}


def build_tag_import_template(project: dict, standard_detail: dict, pbs_nodes: list[dict]) -> bytes:
    workbook = Workbook()

    instruction_sheet = workbook.active
    instruction_sheet.title = INSTRUCTION_SHEET
    instructions = [
        f"项目: {project['code']} {project['name']}",
        "1. 每个 Class 单独一个 sheet，仅支持设备级 TAG 导入。",
        "2. 必填列: tag_no, name, pbs_code。",
        "3. 公共属性在前，当前 Class 独有属性在后，表头中的 [attribute_code] 请勿修改。",
        "4. pbs_code 和枚举属性列已配置 Excel 下拉，可直接从参考数据中选择。",
        "5. 上传后会先生成校验草稿，不会直接写入正式 TAG。",
        "6. 已存在的 tag_no 会进入冲突确认，需在系统里选择更新或跳过。",
    ]
    for index, line in enumerate(instructions, start=1):
        instruction_sheet.cell(row=index, column=1, value=line)
    instruction_sheet.column_dimensions["A"].width = 100

    index_sheet = workbook.create_sheet(INDEX_SHEET)
    index_sheet.append(["sheet_name", "class_code", "class_name", "required_attributes"])
    for cell in index_sheet[1]:
        cell.font = Font(bold=True)

    pbs_sheet = workbook.create_sheet(PBS_REFERENCE_SHEET)
    pbs_sheet.append(["pbs_code", "pbs_name"])
    for cell in pbs_sheet[1]:
        cell.font = Font(bold=True)
    for node in pbs_nodes:
        pbs_sheet.append([node["code"], node["name"]])

    class_sheet = workbook.create_sheet(CLASS_REFERENCE_SHEET)
    class_sheet.append(["class_code", "class_name", "required_attributes"])
    for cell in class_sheet[1]:
        cell.font = Font(bold=True)
    for class_definition in standard_detail.get("classes", []):
        required_attributes = [
            attribute["code"]
            for attribute in class_definition.get("attributes", [])
            if attribute.get("is_required")
        ]
        class_sheet.append(
            [
                class_definition["code"],
                class_definition["name"],
                ", ".join(required_attributes),
            ]
        )

    meta_sheet = workbook.create_sheet(META_SHEET)
    meta_sheet.append(["sheet_name", "class_code", "class_name"])
    meta_sheet.sheet_state = "hidden"

    _add_pbs_named_range(workbook, pbs_nodes)

    common_attributes = [
        {**attribute, "code": _normalize_string(attribute["code"])}
        for attribute in standard_detail.get("common_attributes", [])
    ]
    first_class_code = standard_detail.get("classes", [{}])[0].get("code")
    used_sheet_names = {
        INSTRUCTION_SHEET,
        INDEX_SHEET,
        PBS_REFERENCE_SHEET,
        CLASS_REFERENCE_SHEET,
        META_SHEET,
    }

    for class_definition in standard_detail.get("classes", []):
        sheet_name = _build_unique_sheet_name(
            class_definition["code"],
            class_definition["name"],
            used_sheet_names,
        )
        used_sheet_names.add(sheet_name)
        class_attributes = _build_class_sheet_attributes(common_attributes, class_definition.get("attributes", []))
        headers = list(CLASS_SHEET_REQUIRED_HEADERS) + [
            f"{attribute['name']} [{attribute['code']}]" for attribute in class_attributes
        ]

        meta_sheet.append([sheet_name, class_definition["code"], class_definition["name"]])
        index_sheet.append(
            [
                sheet_name,
                class_definition["code"],
                class_definition["name"],
                ", ".join(
                    attribute["code"]
                    for attribute in class_attributes
                    if attribute.get("is_required")
                ),
            ]
        )

        data_sheet = workbook.create_sheet(sheet_name)
        data_sheet.append(headers)
        data_sheet.append(
            [
                "P-1001" if class_definition["code"] == first_class_code else "",
                "示例设备" if class_definition["code"] == first_class_code else "",
                pbs_nodes[0]["code"] if pbs_nodes and class_definition["code"] == first_class_code else "",
                *["" for _ in class_attributes],
            ]
        )
        for cell in data_sheet[1]:
            cell.font = Font(bold=True)
        data_sheet.freeze_panes = "A2"
        _add_sheet_dropdowns(data_sheet, class_attributes, has_pbs_options=bool(pbs_nodes))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def validate_tag_import_workbook(
    workbook_bytes: bytes,
    *,
    standard_detail: dict,
    pbs_nodes: list[dict],
    existing_tags: list[dict],
) -> dict:
    try:
        row_values = _parse_workbook_rows(workbook_bytes)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
        message = str(error)
        if isinstance(error, (BadZipFile, InvalidFileException, OSError)):
            message = "Uploaded file is not a valid .xlsx workbook"
        raise ValueError(message) from error
    rows = validate_tag_import_rows(
        row_values,
        standard_detail=standard_detail,
        pbs_nodes=pbs_nodes,
        existing_tags=existing_tags,
    )
    return {
        "summary": _summarize_rows(rows),
        "rows": rows,
    }


def validate_tag_import_rows(
    row_values: list[dict],
    *,
    standard_detail: dict,
    pbs_nodes: list[dict],
    existing_tags: list[dict],
) -> list[dict]:
    duplicate_pbs_codes = [
        code
        for code, count in Counter(_normalize_string(node.get("code")) for node in pbs_nodes).items()
        if code and count > 1
    ]
    if duplicate_pbs_codes:
        raise ValueError(
            "PBS codes must be unique before importing TAG data: " + ", ".join(sorted(duplicate_pbs_codes))
        )

    pbs_by_code = {
        _normalize_string(node.get("code")): node
        for node in pbs_nodes
        if _normalize_string(node.get("code"))
    }
    class_by_code = {
        _normalize_string(class_definition.get("code")): class_definition
        for class_definition in standard_detail.get("classes", [])
        if _normalize_string(class_definition.get("code"))
    }
    attribute_definitions = _build_template_attributes(standard_detail)
    attribute_by_code = {attribute["code"]: attribute for attribute in attribute_definitions}
    existing_by_tag_no = {
        _normalize_string(tag.get("tag_no")): tag
        for tag in existing_tags
        if _normalize_string(tag.get("tag_no"))
    }
    duplicate_tag_nos = Counter(
        _normalize_string(row.get("tag_no"))
        for row in row_values
        if _normalize_string(row.get("tag_no"))
    )

    validated_rows = []
    for row in row_values:
        validated_rows.append(
            _validate_row(
                row,
                pbs_by_code=pbs_by_code,
                class_by_code=class_by_code,
                attribute_by_code=attribute_by_code,
                existing_by_tag_no=existing_by_tag_no,
                duplicate_tag_nos=duplicate_tag_nos,
            )
        )
    return validated_rows


def create_tag_import_job_from_upload(project_id: str, filename: str, workbook_bytes: bytes) -> dict:
    project, standard_detail, pbs_nodes, existing_tags = _load_import_context(project_id)
    if not filename.lower().endswith(".xlsx"):
        raise ValueError("Only .xlsx files are supported")
    if len(workbook_bytes) > MAX_UPLOAD_SIZE:
        raise ValueError("Uploaded file exceeds the 10 MB limit")

    validated = validate_tag_import_workbook(
        workbook_bytes,
        standard_detail=standard_detail,
        pbs_nodes=pbs_nodes,
        existing_tags=existing_tags,
    )
    created = _store_tag_import_job(
        project_id=project_id,
        filename=filename,
        summary=validated["summary"],
        rows=validated["rows"],
    )
    return get_tag_import_job_detail(project_id, created["id"])


def get_tag_import_job_detail(
    project_id: str,
    job_id: str,
    *,
    status: str | None = None,
    page: int = 1,
    page_size: int = DEFAULT_PAGE_SIZE,
) -> dict:
    normalized_page_size = max(1, min(page_size, MAX_PAGE_SIZE))
    normalized_page = max(1, page)

    job = fetch_one(
        """
        SELECT id, project_id, filename, summary, status, created_at, updated_at, committed_at
        FROM tag_import_job
        WHERE id = %s
          AND project_id = %s
        """,
        (job_id, project_id),
    )
    if job is None:
        raise ValueError("Tag import job not found")

    status_filter_sql = ""
    status_params: list[Any] = []
    if status:
        status_filter_sql = " AND r.status = %s"
        status_params.append(status)

    total_rows_result = fetch_one(
        f"""
        SELECT COUNT(*)::int AS total_rows
        FROM tag_import_row r
        WHERE r.job_id = %s
        {status_filter_sql}
        """,
        (job_id, *status_params),
    )
    total_rows = total_rows_result["total_rows"] if total_rows_result else 0
    total_pages = max(1, math.ceil(total_rows / normalized_page_size)) if total_rows else 1
    offset = (normalized_page - 1) * normalized_page_size

    with get_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT
                    r.id,
                    r.row_number,
                    r.values,
                    r.normalized_values,
                    r.issues,
                    r.status,
                    r.conflict_action,
                    t.id AS existing_tag_id,
                    t.tag_no AS existing_tag_tag_no,
                    t.name AS existing_tag_name,
                    t.pbs_node_id AS existing_tag_pbs_node_id,
                    t.class_id AS existing_tag_class_id,
                    t.attribute_values AS existing_tag_attribute_values
                FROM tag_import_row r
                LEFT JOIN tag t ON t.id = r.existing_tag_id
                WHERE r.job_id = %s
                {status_filter_sql}
                ORDER BY r.row_number ASC
                LIMIT %s OFFSET %s
                """,
                (job_id, *status_params, normalized_page_size, offset),
            )
            rows = [_serialize_tag_import_row(row) for row in cursor.fetchall()]

    return {
        "job_id": job["id"],
        "filename": job["filename"],
        "summary": job["summary"],
        "rows": rows,
        "page": normalized_page,
        "page_size": normalized_page_size,
        "total_pages": total_pages,
        "status": job["status"],
        "created_at": job["created_at"],
        "updated_at": job["updated_at"],
        "committed_at": job["committed_at"],
    }


def patch_tag_import_row(project_id: str, job_id: str, row_id: str, payload: dict) -> dict:
    _ensure_job_exists(project_id, job_id)

    row_to_patch = fetch_one(
        """
        SELECT id, row_number, values
        FROM tag_import_row
        WHERE id = %s
          AND job_id = %s
        """,
        (row_id, job_id),
    )
    if row_to_patch is None:
        raise ValueError("Tag import row not found")

    row_values = _load_job_row_values(job_id)
    merged_rows = []
    for row in row_values:
        next_values = dict(row["values"])
        next_conflict_action = row.get("conflict_action")
        if row["id"] == row_id:
            next_values.update(payload.get("values", {}))
            if "conflict_action" in payload:
                next_conflict_action = payload.get("conflict_action")
        merged_rows.append(
            {
                "id": row["id"],
                "row_number": row["row_number"],
                "sheet_name": row["values"].get("_sheet_name"),
                "source_row_number": row["values"].get("_source_row_number"),
                **next_values,
                "conflict_action": next_conflict_action,
            }
        )

    _, standard_detail, pbs_nodes, existing_tags = _load_import_context(project_id)
    validated_rows = validate_tag_import_rows(
        merged_rows,
        standard_detail=standard_detail,
        pbs_nodes=pbs_nodes,
        existing_tags=existing_tags,
    )
    for row, merged in zip(validated_rows, merged_rows, strict=False):
        if merged.get("conflict_action") in CONFLICT_ACTIONS:
            row["conflict_action"] = merged["conflict_action"]

    summary = _summarize_rows(validated_rows)
    _replace_job_rows(job_id, summary, validated_rows)

    updated_row = next(row for row in validated_rows if row["id"] == row_id)
    return {
        "job_id": job_id,
        "summary": summary,
        "row": updated_row,
    }


def commit_tag_import_job(project_id: str, job_id: str, conflict_actions: list[dict]) -> dict:
    _ensure_job_exists(project_id, job_id)
    rows = _load_job_rows(job_id)
    if not rows:
        raise ValueError("Tag import job has no rows")

    conflict_action_map = {
        action["row_id"]: action["action"]
        for action in conflict_actions
        if action.get("action") in CONFLICT_ACTIONS
    }

    blocking_errors = [row for row in rows if row["status"] == "error"]
    if blocking_errors:
        raise ValueError("Cannot commit while error rows still exist")

    unresolved_conflicts = []
    for row in rows:
        if row["status"] != "conflict":
            continue
        action = conflict_action_map.get(row["id"]) or row.get("conflict_action")
        if action not in CONFLICT_ACTIONS:
            unresolved_conflicts.append(row["row_number"])

    if unresolved_conflicts:
        raise ValueError(
            "Conflict rows must be resolved before commit: "
            + ", ".join(str(row_number) for row_number in unresolved_conflicts)
        )

    created_count = 0
    updated_count = 0
    skipped_count = 0

    with get_connection() as connection:
        with connection.cursor() as cursor:
            for row in rows:
                normalized = row.get("normalized_values", {})
                payload = (
                    normalized.get("tag_no"),
                    normalized.get("name"),
                    normalized.get("pbs_node_id"),
                    normalized.get("class_id"),
                    Json(normalized.get("attribute_values", {})),
                )

                if row["status"] == "conflict":
                    action = conflict_action_map.get(row["id"]) or row.get("conflict_action")
                    if action == "skip":
                        skipped_count += 1
                        continue

                    cursor.execute(
                        """
                        UPDATE tag
                        SET
                            tag_no = %s,
                            name = %s,
                            pbs_node_id = %s,
                            class_id = %s,
                            attribute_values = %s,
                            updated_at = now()
                        WHERE id = %s
                        """,
                        (*payload, row["existing_tag"]["id"]),
                    )
                    updated_count += 1
                    continue

                cursor.execute(
                    """
                    INSERT INTO tag (project_id, tag_no, name, pbs_node_id, class_id, attribute_values, status)
                    VALUES (%s, %s, %s, %s, %s, %s, 'active')
                    """,
                    (project_id, *payload),
                )
                created_count += 1

            cursor.execute(
                """
                UPDATE tag_import_job
                SET status = 'committed',
                    committed_at = now(),
                    updated_at = now()
                WHERE id = %s
                """,
                (job_id,),
            )
        connection.commit()

    return {
        "job_id": job_id,
        "created_count": created_count,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "failed_count": 0,
        "failures": [],
    }


def _load_import_context(project_id: str) -> tuple[dict, dict, list[dict], list[dict]]:
    project = get_project_detail(project_id)
    if project is None:
        raise ValueError("Project not found")

    standard_id = project.get("reference_attributes", {}).get("standard_id")
    if not isinstance(standard_id, str) or not standard_id:
        raise ValueError("Project is not linked to a standard")

    standard_detail = get_standard_detail(standard_id, include_attributes=True)
    if standard_detail is None:
        raise ValueError("Standard not found")

    pbs_nodes = get_pbs_nodes(project_id)
    existing_tags = get_project_tags(project_id)
    return project, standard_detail, pbs_nodes, existing_tags


def _build_template_attributes(standard_detail: dict) -> list[dict]:
    attribute_by_code: dict[str, dict] = {}
    for attribute in standard_detail.get("common_attributes", []):
        attribute_by_code[_normalize_string(attribute["code"])] = {
            **attribute,
            "code": _normalize_string(attribute["code"]),
        }
    for class_definition in standard_detail.get("classes", []):
        for attribute in class_definition.get("attributes", []):
            code = _normalize_string(attribute["code"])
            attribute_by_code.setdefault(code, {**attribute, "code": code})
    return [attribute_by_code[code] for code in sorted(attribute_by_code)]


def _build_class_sheet_attributes(common_attributes: list[dict], class_attributes: list[dict]) -> list[dict]:
    ordered_attributes: list[dict] = []
    seen_codes: set[str] = set()
    for attribute in [*common_attributes, *class_attributes]:
        code = _normalize_string(attribute["code"])
        if not code or code in seen_codes or code in CLASS_SHEET_REQUIRED_HEADERS:
            continue
        seen_codes.add(code)
        ordered_attributes.append({**attribute, "code": code})
    return ordered_attributes


def _parse_workbook_rows(workbook_bytes: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    if META_SHEET not in workbook.sheetnames:
        raise ValueError(f"Workbook must contain a '{META_SHEET}' sheet")

    meta_sheet = workbook[META_SHEET]
    meta_rows = list(meta_sheet.iter_rows(min_row=2, values_only=True))
    if not meta_rows:
        raise ValueError("Workbook does not contain any class sheet metadata")

    rows = []
    import_row_number = 1
    for meta_row in meta_rows:
        sheet_name = _normalize_string(meta_row[0] if len(meta_row) > 0 else None)
        class_code = _normalize_string(meta_row[1] if len(meta_row) > 1 else None)
        if not sheet_name or not class_code:
            continue
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Workbook is missing class sheet '{sheet_name}'")

        sheet = workbook[sheet_name]
        headers = [_normalize_header(cell.value) for cell in sheet[1]]
        missing_headers = [header for header in CLASS_SHEET_REQUIRED_HEADERS if header not in headers]
        if missing_headers:
            raise ValueError(
                f"Sheet '{sheet_name}' is missing required headers: " + ", ".join(missing_headers)
            )

        for row_number, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
            values = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                if header in CLASS_SHEET_REQUIRED_HEADERS and header in values:
                    continue
                values[header] = row_cells[index].value
            if _row_is_empty(values):
                continue
            rows.append(
                {
                    "id": None,
                    "row_number": import_row_number,
                    "sheet_name": sheet_name,
                    "source_row_number": row_number,
                    **values,
                    "class_code": class_code,
                    "conflict_action": None,
                }
            )
            import_row_number += 1
    return rows


def _validate_row(
    row: dict,
    *,
    pbs_by_code: dict[str, dict],
    class_by_code: dict[str, dict],
    attribute_by_code: dict[str, dict],
    existing_by_tag_no: dict[str, dict],
    duplicate_tag_nos: Counter,
) -> dict:
    issues = []
    values = _normalize_row_values(row)
    tag_no = values["tag_no"]
    name = values["name"]
    pbs_code = values["pbs_code"]
    class_code = values["class_code"]
    selected_class = class_by_code.get(class_code) if class_code else None

    if not tag_no:
        issues.append(_issue("required", "tag_no", "Tag 位号不能为空"))
    if not name:
        issues.append(_issue("required", "name", "名称不能为空"))
    if not pbs_code:
        issues.append(_issue("required", "pbs_code", "PBS 编码不能为空"))
    elif pbs_code not in pbs_by_code:
        issues.append(_issue("pbs_not_found", "pbs_code", "PBS 编码不存在"))

    if not class_code:
        issues.append(_issue("class_missing", "class_code", "模板缺少 Class 元数据"))
    elif selected_class is None:
        issues.append(_issue("class_not_found", "class_code", "Class 编码不存在"))

    if tag_no and duplicate_tag_nos[tag_no] > 1:
        issues.append(_issue("duplicate_in_file", "tag_no", "Excel 文件内存在重复的 Tag 位号"))

    attribute_values: dict[str, Any] = {}
    common_required_codes = {
        attribute["code"]
        for attribute in attribute_by_code.values()
        if attribute.get("standard_id") is not None and attribute.get("is_required")
    }
    allowed_attribute_codes = set(common_required_codes)
    if selected_class is not None:
        allowed_attribute_codes.update(
            _normalize_string(attribute["code"]) for attribute in selected_class.get("attributes", [])
        )

    for key, raw_value in values.items():
        if key in REQUIRED_HEADERS:
            continue
        attribute = attribute_by_code.get(key)
        if attribute is None:
            continue

        if selected_class is not None and attribute.get("standard_id") is None and key not in allowed_attribute_codes:
            if _has_value(raw_value):
                issues.append(_issue("attribute_not_allowed", key, "该属性不属于当前 Class"))
            continue

        if not _has_value(raw_value):
            if attribute.get("is_required") and (
                attribute.get("standard_id") is not None
                or (selected_class is not None and key in allowed_attribute_codes)
            ):
                issues.append(_issue("required", key, f"{attribute['name']} 为必填"))
            continue

        try:
            attribute_values[key] = _coerce_attribute_value(attribute, raw_value)
        except ValueError as error:
            message = str(error.args[0]) if error.args else "属性值无效"
            code = str(error.args[1]) if len(error.args) > 1 else "type_invalid"
            issues.append(_issue(code, key, message))

    existing_tag = existing_by_tag_no.get(tag_no)
    status: Literal["ready", "error", "warning", "conflict"]
    if any(issue["severity"] == "error" for issue in issues):
        status = "error"
    elif existing_tag is not None:
        status = "conflict"
    elif issues:
        status = "warning"
    else:
        status = "ready"

    return {
        "id": row.get("id"),
        "row_number": row["row_number"],
        "sheet_name": row.get("sheet_name"),
        "source_row_number": row.get("source_row_number"),
        "values": {
            "_sheet_name": row.get("sheet_name"),
            "_source_row_number": row.get("source_row_number"),
            **values,
        },
        "normalized_values": {
            "tag_no": tag_no,
            "name": name,
            "pbs_code": pbs_code,
            "pbs_node_id": _string_or_none(pbs_by_code.get(pbs_code, {}).get("id")),
            "class_code": class_code or None,
            "class_id": _string_or_none(selected_class.get("id")) if selected_class is not None else None,
            "attribute_values": attribute_values,
        },
        "issues": issues,
        "status": status,
        "existing_tag": existing_tag,
        "conflict_action": row.get("conflict_action"),
    }


def _summarize_rows(rows: list[dict]) -> dict:
    summary = {
        "total_rows": len(rows),
        "ready_rows": 0,
        "error_rows": 0,
        "warning_rows": 0,
        "conflict_rows": 0,
        "resolved_conflict_rows": 0,
        "can_commit": False,
    }
    for row in rows:
        summary[f"{row['status']}_rows"] += 1
        if row["status"] == "conflict" and row.get("conflict_action") in CONFLICT_ACTIONS:
            summary["resolved_conflict_rows"] += 1
    summary["can_commit"] = (
        summary["total_rows"] > 0
        and summary["error_rows"] == 0
        and summary["resolved_conflict_rows"] == summary["conflict_rows"]
    )
    return summary


def _store_tag_import_job(*, project_id: str, filename: str, summary: dict, rows: list[dict]) -> dict:
    with get_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO tag_import_job (project_id, filename, summary, status)
                VALUES (%s, %s, %s, 'validated')
                RETURNING id
                """,
                (project_id, filename, Json(summary)),
            )
            job = cursor.fetchone()
            for row in rows:
                cursor.execute(
                    """
                    INSERT INTO tag_import_row (
                        job_id,
                        row_number,
                        values,
                        normalized_values,
                        issues,
                        status,
                        existing_tag_id,
                        conflict_action
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (
                        job["id"],
                        row["row_number"],
                        Json(row["values"]),
                        Json(row["normalized_values"]),
                        Json(row["issues"]),
                        row["status"],
                        row["existing_tag"]["id"] if row["existing_tag"] else None,
                        row.get("conflict_action"),
                    ),
                )
                created_row = cursor.fetchone()
                row["id"] = created_row["id"]
        connection.commit()
    return job


def _replace_job_rows(job_id: str, summary: dict, rows: list[dict]) -> None:
    with get_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                UPDATE tag_import_job
                SET summary = %s,
                    updated_at = now()
                WHERE id = %s
                """,
                (Json(summary), job_id),
            )
            for row in rows:
                cursor.execute(
                    """
                    UPDATE tag_import_row
                    SET values = %s,
                        normalized_values = %s,
                        issues = %s,
                        status = %s,
                        existing_tag_id = %s,
                        conflict_action = %s,
                        updated_at = now()
                    WHERE id = %s
                      AND job_id = %s
                    """,
                    (
                        Json(row["values"]),
                        Json(row["normalized_values"]),
                        Json(row["issues"]),
                        row["status"],
                        row["existing_tag"]["id"] if row["existing_tag"] else None,
                        row.get("conflict_action"),
                        row["id"],
                        job_id,
                    ),
                )
        connection.commit()


def _load_job_row_values(job_id: str) -> list[dict]:
    with get_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, row_number, values, conflict_action
                FROM tag_import_row
                WHERE job_id = %s
                ORDER BY row_number ASC
                """,
                (job_id,),
            )
            return list(cursor.fetchall())


def _load_job_rows(job_id: str) -> list[dict]:
    with get_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    r.id,
                    r.row_number,
                    r.values,
                    r.normalized_values,
                    r.issues,
                    r.status,
                    r.conflict_action,
                    t.id AS existing_tag_id,
                    t.tag_no AS existing_tag_tag_no,
                    t.name AS existing_tag_name,
                    t.pbs_node_id AS existing_tag_pbs_node_id,
                    t.class_id AS existing_tag_class_id,
                    t.attribute_values AS existing_tag_attribute_values
                FROM tag_import_row r
                LEFT JOIN tag t ON t.id = r.existing_tag_id
                WHERE r.job_id = %s
                ORDER BY r.row_number ASC
                """,
                (job_id,),
            )
            return [_serialize_tag_import_row(row) for row in cursor.fetchall()]


def _ensure_job_exists(project_id: str, job_id: str) -> None:
    if fetch_one(
        """
        SELECT id
        FROM tag_import_job
        WHERE id = %s
          AND project_id = %s
        """,
        (job_id, project_id),
    ) is None:
        raise ValueError("Tag import job not found")


def _serialize_tag_import_row(row: dict) -> dict:
    existing_tag = None
    if row.get("existing_tag_id") is not None:
        existing_tag = {
            "id": row["existing_tag_id"],
            "tag_no": row["existing_tag_tag_no"],
            "name": row["existing_tag_name"],
            "pbs_node_id": row["existing_tag_pbs_node_id"],
            "class_id": row["existing_tag_class_id"],
            "attribute_values": row["existing_tag_attribute_values"],
        }
    values = dict(row["values"])
    sheet_name = values.pop("_sheet_name", None)
    source_row_number = values.pop("_source_row_number", None)
    return {
        "id": row["id"],
        "row_number": row["row_number"],
        "sheet_name": sheet_name,
        "source_row_number": source_row_number,
        "values": values,
        "normalized_values": row["normalized_values"],
        "issues": row["issues"],
        "status": row["status"],
        "existing_tag": existing_tag,
        "conflict_action": row.get("conflict_action"),
    }


def _normalize_header(value: Any) -> str:
    text = _normalize_string(value)
    if not text:
        return ""
    if text in REQUIRED_HEADERS or text in CLASS_SHEET_REQUIRED_HEADERS:
        return text
    if "[" in text and text.endswith("]"):
        return _normalize_string(text.rsplit("[", 1)[1][:-1])
    return text


def _normalize_row_values(row: dict) -> dict:
    normalized = {
        "tag_no": _normalize_string(row.get("tag_no")),
        "name": _normalize_string(row.get("name")),
        "pbs_code": _normalize_string(row.get("pbs_code")),
        "class_code": _normalize_string(row.get("class_code")),
    }
    for key, value in row.items():
        if key in {"id", "row_number", "conflict_action", "sheet_name", "source_row_number", *REQUIRED_HEADERS}:
            continue
        if str(key).startswith("_"):
            continue
        normalized[_normalize_string(key)] = value
    return normalized


def _normalize_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _string_or_none(value: Any) -> str | None:
    if value is None:
        return None
    return str(value)


def _row_is_empty(values: dict) -> bool:
    return not any(_has_value(value) for value in values.values())


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def _issue(code: str, field: str, message: str, *, severity: Literal["error", "warning"] = "error") -> dict:
    return {
        "code": code,
        "field": field,
        "message": message,
        "severity": severity,
    }


def _coerce_attribute_value(attribute: dict, raw_value: Any) -> Any:
    value_type = attribute["value_type"]
    if value_type == "string":
        return _normalize_string(raw_value)
    if value_type == "number":
        try:
            return float(raw_value)
        except (TypeError, ValueError) as error:
            raise ValueError(f"{attribute['name']} 必须是数字", "type_invalid") from error
    if value_type == "integer":
        try:
            string_value = _normalize_string(raw_value)
            if "." in string_value:
                raise ValueError
            return int(string_value)
        except (TypeError, ValueError) as error:
            raise ValueError(f"{attribute['name']} 必须是整数", "type_invalid") from error
    if value_type == "boolean":
        if isinstance(raw_value, bool):
            return raw_value
        normalized = _normalize_string(raw_value).lower()
        if normalized in {"true", "1", "yes", "y", "是"}:
            return True
        if normalized in {"false", "0", "no", "n", "否"}:
            return False
        raise ValueError(f"{attribute['name']} 必须是布尔值", "type_invalid")
    if value_type == "date":
        if isinstance(raw_value, datetime):
            return raw_value.date().isoformat()
        if isinstance(raw_value, date):
            return raw_value.isoformat()
        normalized = _normalize_string(raw_value)
        for pattern in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(normalized, pattern).date().isoformat()
            except ValueError:
                continue
        raise ValueError(f"{attribute['name']} 必须是日期", "type_invalid")
    if value_type == "enum":
        normalized = _normalize_string(raw_value)
        allowed_values = {_normalize_string(option) for option in attribute.get("enum_options", [])}
        if normalized not in allowed_values:
            raise ValueError(
                f"{attribute['name']} 必须是以下值之一: {', '.join(sorted(allowed_values))}",
                "enum_invalid",
            )
        return normalized
    if value_type == "json":
        if isinstance(raw_value, (dict, list)):
            return raw_value
        try:
            return json.loads(_normalize_string(raw_value))
        except JSONDecodeError as error:
            raise ValueError(f"{attribute['name']} 必须是合法 JSON", "type_invalid") from error
    return raw_value


def _sanitize_sheet_name(code: str, name: str) -> str:
    raw = f"{_normalize_string(code)}_{_normalize_string(name)}".strip("_") or "Class"
    sanitized = "".join("_" if char in "\\/?*[]:" else char for char in raw)
    sanitized = " ".join(sanitized.split())
    return sanitized[:31] or "Class"


def _build_unique_sheet_name(code: str, name: str, used_names: set[str]) -> str:
    base_name = _sanitize_sheet_name(code, name)
    if base_name not in used_names:
        return base_name

    suffix = 2
    while True:
        candidate = f"{base_name[: max(0, 31 - len(str(suffix)) - 1)]}_{suffix}"
        if candidate not in used_names:
            return candidate
        suffix += 1


def _add_pbs_named_range(workbook: Workbook, pbs_nodes: list[dict]) -> None:
    if not pbs_nodes:
        return
    reference = f"{quote_sheetname(PBS_REFERENCE_SHEET)}!$A$2:$A${len(pbs_nodes) + 1}"
    workbook.defined_names.add(DefinedName(PBS_CODES_RANGE_NAME, attr_text=reference))


def _add_sheet_dropdowns(data_sheet, class_attributes: list[dict], *, has_pbs_options: bool) -> None:
    if has_pbs_options:
        pbs_validation = DataValidation(type="list", formula1=f"={PBS_CODES_RANGE_NAME}", allow_blank=False)
        pbs_validation.prompt = "请选择项目内已有 PBS 节点编码"
        pbs_validation.error = "请选择 PBS参考 中的编码"
        data_sheet.add_data_validation(pbs_validation)
        pbs_validation.add(f"C2:C{TEMPLATE_MAX_ROWS}")

    for column_index, attribute in enumerate(class_attributes, start=4):
        if attribute.get("value_type") != "enum" or not attribute.get("enum_options"):
            continue
        options = [_normalize_string(option) for option in attribute.get("enum_options", []) if _normalize_string(option)]
        if not options:
            continue
        validation = DataValidation(
            type="list",
            formula1='"' + ",".join(options).replace('"', '""') + '"',
            allow_blank=not attribute.get("is_required", False),
        )
        validation.prompt = f"请选择 {attribute['name']} 的合法值"
        validation.error = f"{attribute['name']} 只能选择预设枚举值"
        data_sheet.add_data_validation(validation)
        column_letter = get_column_letter(column_index)
        validation.add(f"{column_letter}2:{column_letter}{TEMPLATE_MAX_ROWS}")
