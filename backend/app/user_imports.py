from __future__ import annotations

import io
import math
import re
from collections import Counter
from typing import Any, Literal
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
from psycopg.types.json import Json

from .auth_repository import list_roles_by_scope, list_users
from .db import fetch_one, get_connection
from .i18n import DEFAULT_LOCALE, translate
from .security import hash_password


INSTRUCTION_SHEET = "说明"
IMPORT_SHEET = "用户导入"
EXPORT_SHEET = "用户导出"
ROLE_REFERENCE_SHEET = "系统角色参考"
IMPORT_SHEET_BY_LOCALE = {
    "zh-CN": IMPORT_SHEET,
    "en-US": "User Import",
}
ROLE_REFERENCE_SHEET_BY_LOCALE = {
    "zh-CN": ROLE_REFERENCE_SHEET,
    "en-US": "Role Reference",
}
IMPORT_HEADER_CODES = ["username", "display_name", "email", "status", "password", "system_role_codes"]
IMPORT_HEADER_LABEL_CODES = [
    "userImportHeaderUsername",
    "userImportHeaderDisplayName",
    "userImportHeaderEmail",
    "userImportHeaderStatus",
    "userImportHeaderPassword",
    "userImportHeaderSystemRoleCodes",
]
LOCALIZED_IMPORT_HEADERS = {
    translate(code, locale): field
    for locale in ("zh-CN", "en-US")
    for field, code in zip(IMPORT_HEADER_CODES, IMPORT_HEADER_LABEL_CODES, strict=True)
}
DEFAULT_PAGE_SIZE = 50
MAX_PAGE_SIZE = 200
MAX_UPLOAD_SIZE = 10 * 1024 * 1024
MASKED_PASSWORD_DISPLAY = "••••••••"
VALID_STATUSES = {"active", "disabled"}
ROW_STATUS_VALUES = {"ready", "error", "warning"}
ROW_ACTION_VALUES = {"create", "update", "skip"}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def build_user_import_template(system_roles: list[dict] | None = None, *, locale: str | None = None) -> bytes:
    roles = _normalize_system_roles(system_roles if system_roles is not None else list_roles_by_scope("system"))
    effective_locale = locale or DEFAULT_LOCALE
    should_localize = locale is not None

    workbook = Workbook()
    instruction_sheet = workbook.active
    instruction_sheet.title = INSTRUCTION_SHEET
    instructions = [
        "1. 仅支持 .xlsx，上传后会先生成导入预览，不会立即写入正式用户表。",
        "2. 必填列: username, display_name。",
        "3. 新用户 password 必填且至少 8 位；已有用户 password 留空表示不修改密码。",
        "4. status 仅支持 active / disabled，留空默认 active。",
        "5. system_role_codes 多个角色可用英文逗号或分号分隔；留空表示清空系统角色。",
        "6. 以 username 匹配已有用户，用户名大小写不敏感。",
    ]
    for index, line in enumerate(instructions, start=1):
        instruction_sheet.cell(row=index, column=1, value=line)
    instruction_sheet.column_dimensions["A"].width = 110

    import_sheet = workbook.create_sheet(IMPORT_SHEET_BY_LOCALE.get(effective_locale, IMPORT_SHEET) if should_localize else IMPORT_SHEET)
    import_sheet.append([translate(code, effective_locale) for code in IMPORT_HEADER_LABEL_CODES] if should_localize else IMPORT_HEADER_CODES)
    import_sheet.append(["new.user", "New User" if effective_locale == "en-US" else "新用户", "new.user@example.test", "active", "ChangeMe123", "project_creator"])
    for cell in import_sheet[1]:
        cell.font = Font(bold=True)
    import_sheet.freeze_panes = "A2"
    for column, width in {"A": 20, "B": 20, "C": 28, "D": 14, "E": 18, "F": 36}.items():
        import_sheet.column_dimensions[column].width = width

    role_sheet = workbook.create_sheet(ROLE_REFERENCE_SHEET_BY_LOCALE.get(effective_locale, ROLE_REFERENCE_SHEET) if should_localize else ROLE_REFERENCE_SHEET)
    role_sheet.append([translate("roleCode", effective_locale), translate("roleName", effective_locale), translate("permissions", effective_locale)] if should_localize else ["code", "name", "permissions"])
    for cell in role_sheet[1]:
        cell.font = Font(bold=True)
    for role in roles:
        role_sheet.append([role["code"], role["name"], " / ".join(role.get("permissions") or [])])
    role_sheet.freeze_panes = "A2"
    role_sheet.column_dimensions["A"].width = 24
    role_sheet.column_dimensions["B"].width = 28
    role_sheet.column_dimensions["C"].width = 64

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_user_export_workbook(users: list[dict] | None = None, system_roles: list[dict] | None = None) -> bytes:
    exported_users = users if users is not None else list_users()
    roles = _normalize_system_roles(system_roles if system_roles is not None else list_roles_by_scope("system"))

    workbook = Workbook()
    export_sheet = workbook.active
    export_sheet.title = EXPORT_SHEET
    export_sheet.append(
        [
            "username",
            "display_name",
            "email",
            "status",
            "system_role_codes",
            "system_role_names",
            "last_login_at",
            "created_at",
            "updated_at",
            "password",
        ]
    )
    for cell in export_sheet[1]:
        cell.font = Font(bold=True)

    for user in exported_users:
        export_sheet.append(
            [
                user["username"],
                user["display_name"],
                user.get("email"),
                user["status"],
                ",".join(user.get("role_codes") or []),
                ",".join(user.get("role_names") or []),
                _string_value(user.get("last_login_at")),
                _string_value(user.get("created_at")),
                _string_value(user.get("updated_at")),
                "",
            ]
        )

    export_sheet.freeze_panes = "A2"
    for column, width in {"A": 20, "B": 20, "C": 28, "D": 14, "E": 30, "F": 30, "G": 24, "H": 24, "I": 24, "J": 18}.items():
        export_sheet.column_dimensions[column].width = width

    role_sheet = workbook.create_sheet(ROLE_REFERENCE_SHEET)
    role_sheet.append(["code", "name", "permissions"])
    for cell in role_sheet[1]:
        cell.font = Font(bold=True)
    for role in roles:
        role_sheet.append([role["code"], role["name"], " / ".join(role.get("permissions") or [])])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def create_user_import_job_from_upload(filename: str, workbook_bytes: bytes, *, allow_role_management: bool) -> dict:
    if not filename.lower().endswith(".xlsx"):
        raise ValueError("Only .xlsx files are supported")
    if len(workbook_bytes) > MAX_UPLOAD_SIZE:
        raise ValueError("Uploaded file exceeds the 10 MB limit")

    validated = validate_user_import_workbook(
        workbook_bytes,
        existing_users=list_users(),
        system_roles=list_roles_by_scope("system"),
        allow_role_management=allow_role_management,
    )
    created = _store_user_import_job(filename=filename, summary=validated["summary"], rows=validated["rows"])
    return get_user_import_job_detail(str(created["id"]))


def get_user_import_job_detail(
    job_id: str,
    *,
    status: str | None = None,
    page: int = 1,
    page_size: int = DEFAULT_PAGE_SIZE,
) -> dict:
    normalized_page_size = max(1, min(page_size, MAX_PAGE_SIZE))
    normalized_page = max(1, page)
    filter_kind = _normalize_filter(status)

    job = fetch_one(
        """
        SELECT id, filename, summary, status, created_at, updated_at, committed_at
        FROM user_import_job
        WHERE id = %s
        """,
        (job_id,),
    )
    if job is None:
        raise ValueError("User import job not found")

    filter_sql = ""
    filter_params: list[Any] = []
    if filter_kind is not None:
        column_name = "action" if filter_kind in ROW_ACTION_VALUES else "status"
        filter_sql = f" AND r.{column_name} = %s"
        filter_params.append(filter_kind)

    total_rows_result = fetch_one(
        f"""
        SELECT COUNT(*)::int AS total_rows
        FROM user_import_row r
        WHERE r.job_id = %s
        {filter_sql}
        """,
        (job_id, *filter_params),
    )
    total_rows = total_rows_result["total_rows"] if total_rows_result else 0
    total_pages = max(1, math.ceil(total_rows / normalized_page_size)) if total_rows else 1
    offset = (normalized_page - 1) * normalized_page_size

    with get_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT
                    r.id,
                    r.row_number,
                    r.values,
                    r.normalized_values,
                    r.issues,
                    r.status,
                    r.action,
                    u.id AS existing_user_id,
                    u.username AS existing_user_username,
                    u.email AS existing_user_email,
                    u.display_name AS existing_user_display_name,
                    u.status AS existing_user_status
                FROM user_import_row r
                LEFT JOIN user_account u ON u.id = r.existing_user_id
                WHERE r.job_id = %s
                {filter_sql}
                ORDER BY r.row_number ASC
                LIMIT %s OFFSET %s
                """,
                (job_id, *filter_params, normalized_page_size, offset),
            )
            rows = [_serialize_user_import_row(row) for row in cursor.fetchall()]

    return {
        "job_id": str(job["id"]),
        "filename": job["filename"],
        "summary": job["summary"],
        "rows": rows,
        "page": normalized_page,
        "page_size": normalized_page_size,
        "total_pages": total_pages,
        "status": job["status"],
        "created_at": job["created_at"],
        "updated_at": job["updated_at"],
        "committed_at": job["committed_at"],
    }


def patch_user_import_row(job_id: str, row_id: str, payload: dict, *, allow_role_management: bool) -> dict:
    _ensure_user_import_job_exists(job_id)
    row_to_patch = fetch_one(
        """
        SELECT id
        FROM user_import_row
        WHERE id = %s
          AND job_id = %s
        """,
        (row_id, job_id),
    )
    if row_to_patch is None:
        raise ValueError("User import row not found")

    row_values = _load_user_import_row_values(job_id)
    merged_rows: list[dict] = []
    for row in row_values:
        next_values = dict(row["values"])
        next_password_hash = _password_hash_or_none(row.get("normalized_values"))
        updates = payload.get("values", {}) if str(row["id"]) == row_id else {}
        for key, value in updates.items():
            next_values[key] = value
            if key == "password":
                if _trim_text(value) == "":
                    next_password_hash = None
        merged_row = {
            "id": str(row["id"]),
            "row_number": row["row_number"],
            **next_values,
        }
        if next_password_hash:
            merged_row["_password_hash"] = next_password_hash
        merged_rows.append(merged_row)

    validated_rows = validate_user_import_rows(
        merged_rows,
        existing_users=list_users(),
        system_roles=list_roles_by_scope("system"),
        allow_role_management=allow_role_management,
    )
    summary = _summarize_user_rows(validated_rows)
    _replace_user_import_rows(job_id, summary, validated_rows)
    updated_row = next(row for row in validated_rows if row["id"] == row_id)
    return {"job_id": job_id, "summary": summary, "row": _sanitize_user_import_row_for_response(updated_row)}


def commit_user_import_job(job_id: str, *, granted_by: str, allow_role_management: bool) -> dict:
    _ensure_user_import_job_exists(job_id)
    stored_rows = _load_user_import_row_values(job_id)
    if not stored_rows:
        raise ValueError("User import job has no rows")

    validation_input: list[dict] = []
    for row in stored_rows:
        next_row = {"id": str(row["id"]), "row_number": row["row_number"], **dict(row["values"])}
        password_hash = _password_hash_or_none(row.get("normalized_values"))
        if password_hash:
            next_row["_password_hash"] = password_hash
        validation_input.append(next_row)

    validated_rows = validate_user_import_rows(
        validation_input,
        existing_users=list_users(),
        system_roles=list_roles_by_scope("system"),
        allow_role_management=allow_role_management,
    )
    summary = _summarize_user_rows(validated_rows)
    _replace_user_import_rows(job_id, summary, validated_rows)

    if summary["error_rows"] > 0:
        raise ValueError("Cannot commit while error rows still exist")

    created_count = 0
    updated_count = 0
    skipped_count = 0
    password_change_count = 0
    disabled_count = 0

    with get_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, code
                FROM role_definition
                WHERE scope_kind = 'system'
                  AND status = 'active'
                """
            )
            system_role_ids = {row["code"]: row["id"] for row in cursor.fetchall()}

            for row in validated_rows:
                normalized = row["normalized_values"]
                if row["action"] == "skip":
                    skipped_count += 1
                    continue

                if row["action"] == "create":
                    if not normalized["password_hash"]:
                        raise ValueError("Cannot create a new user without a password")
                    cursor.execute(
                        """
                        INSERT INTO user_account (username, email, display_name, password_hash, status, metadata)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        RETURNING id
                        """,
                        (
                            normalized["username"],
                            normalized["email"],
                            normalized["display_name"],
                            normalized["password_hash"],
                            normalized["status"],
                            Json({}),
                        ),
                    )
                    created_user = cursor.fetchone()
                    if created_user is None:
                        raise ValueError("User was not created")
                    user_id = str(created_user["id"])
                    created_count += 1
                else:
                    existing_user_id = row["existing_user"]["id"]
                    cursor.execute(
                        """
                        SELECT id, password_hash
                        FROM user_account
                        WHERE id = %s
                        """,
                        (existing_user_id,),
                    )
                    current_user = cursor.fetchone()
                    if current_user is None:
                        raise ValueError("User not found")
                    next_password_hash = normalized["password_hash"] or current_user["password_hash"]
                    cursor.execute(
                        """
                        UPDATE user_account
                        SET
                            email = %s,
                            display_name = %s,
                            password_hash = %s,
                            status = %s,
                            updated_at = now()
                        WHERE id = %s
                        RETURNING id
                        """,
                        (
                            normalized["email"],
                            normalized["display_name"],
                            next_password_hash,
                            normalized["status"],
                            existing_user_id,
                        ),
                    )
                    updated_user = cursor.fetchone()
                    if updated_user is None:
                        raise ValueError("User not found")
                    user_id = existing_user_id
                    updated_count += 1

                if normalized["password_supplied"]:
                    password_change_count += 1

                if normalized["status"] == "disabled":
                    cursor.execute(
                        """
                        UPDATE user_session
                        SET revoked_at = now(),
                            updated_at = now()
                        WHERE user_id = %s
                          AND revoked_at IS NULL
                        """,
                        (user_id,),
                    )
                    disabled_count += 1

                if allow_role_management:
                    cursor.execute(
                        """
                        DELETE FROM user_role_assignment
                        WHERE user_id = %s
                          AND role_id IN (
                              SELECT id
                              FROM role_definition
                              WHERE scope_kind = 'system'
                          )
                        """,
                        (user_id,),
                    )
                    for role_code in normalized["system_role_codes"]:
                        role_id = system_role_ids.get(role_code)
                        if role_id is None:
                            raise ValueError(f"Unknown system role code: {role_code}")
                        cursor.execute(
                            """
                            INSERT INTO user_role_assignment (user_id, role_id, scope_id, granted_by)
                            VALUES (%s, %s, NULL, %s)
                            ON CONFLICT DO NOTHING
                            """,
                            (user_id, role_id, granted_by),
                        )

            cursor.execute(
                """
                UPDATE user_import_job
                SET summary = %s,
                    status = 'committed',
                    committed_at = now(),
                    updated_at = now()
                WHERE id = %s
                """,
                (Json(summary), job_id),
            )
        connection.commit()

    return {
        "job_id": job_id,
        "created_count": created_count,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "failed_count": 0,
        "failures": [],
        "password_change_count": password_change_count,
        "disabled_count": disabled_count,
    }


def validate_user_import_workbook(
    workbook_bytes: bytes,
    *,
    existing_users: list[dict],
    system_roles: list[dict],
    allow_role_management: bool,
) -> dict:
    try:
        row_values = _parse_user_import_workbook(workbook_bytes)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as error:
        message = str(error)
        if isinstance(error, (BadZipFile, InvalidFileException, OSError)):
            message = "Uploaded file is not a valid .xlsx workbook"
        raise ValueError(message) from error

    rows = validate_user_import_rows(
        row_values,
        existing_users=existing_users,
        system_roles=system_roles,
        allow_role_management=allow_role_management,
    )
    return {"summary": _summarize_user_rows(rows), "rows": rows}


def validate_user_import_rows(
    row_values: list[dict],
    *,
    existing_users: list[dict],
    system_roles: list[dict],
    allow_role_management: bool,
) -> list[dict]:
    users_by_username = {
        _username_key(user.get("username")): user
        for user in existing_users
        if _username_key(user.get("username"))
    }
    users_by_email = {
        _normalize_email(user.get("email")): user
        for user in existing_users
        if _normalize_email(user.get("email"))
    }
    duplicate_usernames = Counter(
        _username_key(row.get("username"))
        for row in row_values
        if _username_key(row.get("username"))
    )
    duplicate_emails = Counter(
        _normalize_email(row.get("email"))
        for row in row_values
        if _normalize_email(row.get("email"))
    )
    valid_role_codes = {role["code"] for role in _normalize_system_roles(system_roles)}

    validated_rows: list[dict] = []
    for row in row_values:
        validated_rows.append(
            _validate_user_row(
                row,
                users_by_username=users_by_username,
                users_by_email=users_by_email,
                duplicate_usernames=duplicate_usernames,
                duplicate_emails=duplicate_emails,
                valid_role_codes=valid_role_codes,
                allow_role_management=allow_role_management,
            )
        )
    return validated_rows


def _parse_user_import_workbook(workbook_bytes: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    import_sheet_name = next((sheet_name for sheet_name in IMPORT_SHEET_BY_LOCALE.values() if sheet_name in workbook.sheetnames), None)
    if import_sheet_name is None:
        raise ValueError(f"Workbook must contain a '{IMPORT_SHEET}' sheet")

    sheet = workbook[import_sheet_name]
    headers = [_normalize_header(cell.value) for cell in sheet[1]]
    expected_headers = IMPORT_HEADER_CODES
    missing_headers = [header for header in expected_headers if header not in headers]
    if missing_headers:
        raise ValueError("Workbook is missing required headers: " + ", ".join(missing_headers))

    rows: list[dict] = []
    import_row_number = 1
    for _source_row_number, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        values: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            values[header] = row_cells[index].value
        if _row_is_empty(values):
            continue
        rows.append({"id": None, "row_number": import_row_number, **values})
        import_row_number += 1
    return rows


def _validate_user_row(
    row: dict,
    *,
    users_by_username: dict[str, dict],
    users_by_email: dict[str, dict],
    duplicate_usernames: Counter,
    duplicate_emails: Counter,
    valid_role_codes: set[str],
    allow_role_management: bool,
) -> dict:
    username = _trim_text(row.get("username"))
    username_key = _username_key(username)
    display_name = _trim_text(row.get("display_name"))
    email = _normalize_email(row.get("email"))
    status = _normalize_status(row.get("status"))
    raw_password = _normalize_password(row.get("password"))
    preserved_password_hash = _trim_text(row.get("_password_hash"))
    system_role_codes = _normalize_role_codes(row.get("system_role_codes"))

    existing_user = users_by_username.get(username_key)
    issues: list[dict] = []
    warning_issues: list[dict] = []

    if not username:
        issues.append(_issue("required", "username", "用户名不能为空"))
    if not display_name:
        issues.append(_issue("required", "display_name", "显示名称不能为空"))
    if username_key and duplicate_usernames[username_key] > 1:
        issues.append(_issue("duplicate_username_in_file", "username", "Excel 文件内存在重复用户名"))

    if email and not EMAIL_RE.match(email):
        issues.append(_issue("invalid_email", "email", "邮箱格式不正确"))
    if email and duplicate_emails[email] > 1:
        issues.append(_issue("duplicate_email_in_file", "email", "Excel 文件内存在重复邮箱"))
    if email:
        email_owner = users_by_email.get(email)
        if email_owner is not None and (existing_user is None or str(email_owner["id"]) != str(existing_user["id"])):
            issues.append(_issue("duplicate_email", "email", "邮箱已被其他用户占用"))

    if status not in VALID_STATUSES:
        issues.append(_issue("invalid_status", "status", "状态仅支持 active 或 disabled"))

    password_hash = preserved_password_hash or None
    password_supplied = bool(password_hash)
    if raw_password and raw_password != MASKED_PASSWORD_DISPLAY:
        if len(raw_password) < 8:
            issues.append(_issue("password_too_short", "password", "密码至少需要 8 位"))
        else:
            password_hash = hash_password(raw_password)
            password_supplied = True
    elif raw_password == "":
        password_hash = preserved_password_hash or None
        password_supplied = bool(password_hash)

    action: Literal["create", "update", "skip"] = "update" if existing_user else "create"
    if action == "create" and not password_supplied:
        issues.append(_issue("required_password", "password", "新用户必须提供至少 8 位密码"))

    unknown_role_codes = [code for code in system_role_codes if code not in valid_role_codes]
    if unknown_role_codes:
        issues.append(
            _issue(
                "unknown_role_code",
                "system_role_codes",
                "存在未知系统角色: " + ", ".join(unknown_role_codes),
            )
        )
    if system_role_codes and not allow_role_management:
        issues.append(_issue("role_management_forbidden", "system_role_codes", "当前账号不能批量修改系统角色"))

    existing_role_codes = sorted(existing_user.get("role_codes") or []) if existing_user else []
    existing_display_name = _trim_text(existing_user.get("display_name")) if existing_user else ""
    existing_email = _normalize_email(existing_user.get("email")) if existing_user else None
    existing_status = existing_user.get("status") if existing_user else None
    role_changed = allow_role_management and sorted(system_role_codes) != existing_role_codes
    profile_changed = (
        action == "create"
        or display_name != existing_display_name
        or email != existing_email
        or status != existing_status
        or password_supplied
    )
    if action == "update" and not profile_changed and not role_changed:
        action = "skip"

    if action == "update" and password_supplied:
        warning_issues.append(_issue("password_reset", "password", "将重置该用户密码", severity="warning"))
    if action == "update" and status == "disabled" and existing_status != "disabled":
        warning_issues.append(_issue("will_disable_user", "status", "将停用账号并撤销现有会话", severity="warning"))
    if action == "update" and role_changed:
        warning_issues.append(_issue("system_roles_changed", "system_role_codes", "系统角色将被覆盖", severity="warning"))

    all_issues = [*issues, *warning_issues]
    if any(issue["severity"] == "error" for issue in all_issues):
        row_status: Literal["ready", "error", "warning"] = "error"
    elif warning_issues:
        row_status = "warning"
    else:
        row_status = "ready"

    return {
        "id": row.get("id"),
        "row_number": row["row_number"],
        "action": action,
        "status": row_status,
        "values": {
            "username": username,
            "display_name": display_name,
            "email": email,
            "status": status,
            "password": MASKED_PASSWORD_DISPLAY if password_supplied else "",
            "system_role_codes": ",".join(system_role_codes),
        },
        "normalized_values": {
            "username": username,
            "email": email,
            "display_name": display_name,
            "status": status,
            "password_hash": password_hash,
            "password_supplied": password_supplied,
            "system_role_codes": system_role_codes,
        },
        "issues": all_issues,
        "existing_user": _serialize_existing_user(existing_user),
    }


def _summarize_user_rows(rows: list[dict]) -> dict:
    summary = {
        "total_rows": len(rows),
        "create_rows": 0,
        "update_rows": 0,
        "skip_rows": 0,
        "ready_rows": 0,
        "error_rows": 0,
        "warning_rows": 0,
        "can_commit": False,
    }
    for row in rows:
        summary[f"{row['action']}_rows"] += 1
        summary[f"{row['status']}_rows"] += 1
    summary["can_commit"] = summary["total_rows"] > 0 and summary["error_rows"] == 0
    return summary


def _store_user_import_job(*, filename: str, summary: dict, rows: list[dict]) -> dict:
    with get_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO user_import_job (filename, summary, status)
                VALUES (%s, %s, 'validated')
                RETURNING id
                """,
                (filename, Json(summary)),
            )
            job = cursor.fetchone()
            for row in rows:
                cursor.execute(
                    """
                    INSERT INTO user_import_row (
                        job_id,
                        row_number,
                        values,
                        normalized_values,
                        issues,
                        status,
                        action,
                        existing_user_id
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (
                        job["id"],
                        row["row_number"],
                        Json(row["values"]),
                        Json(row["normalized_values"]),
                        Json(row["issues"]),
                        row["status"],
                        row["action"],
                        row["existing_user"]["id"] if row["existing_user"] else None,
                    ),
                )
                created_row = cursor.fetchone()
                row["id"] = str(created_row["id"])
        connection.commit()
    return job


def _replace_user_import_rows(job_id: str, summary: dict, rows: list[dict]) -> None:
    with get_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                UPDATE user_import_job
                SET summary = %s,
                    updated_at = now()
                WHERE id = %s
                """,
                (Json(summary), job_id),
            )
            for row in rows:
                cursor.execute(
                    """
                    UPDATE user_import_row
                    SET values = %s,
                        normalized_values = %s,
                        issues = %s,
                        status = %s,
                        action = %s,
                        existing_user_id = %s,
                        updated_at = now()
                    WHERE id = %s
                      AND job_id = %s
                    """,
                    (
                        Json(row["values"]),
                        Json(row["normalized_values"]),
                        Json(row["issues"]),
                        row["status"],
                        row["action"],
                        row["existing_user"]["id"] if row["existing_user"] else None,
                        row["id"],
                        job_id,
                    ),
                )
        connection.commit()


def _load_user_import_row_values(job_id: str) -> list[dict]:
    with get_connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT id, row_number, values, normalized_values
                FROM user_import_row
                WHERE job_id = %s
                ORDER BY row_number ASC
                """,
                (job_id,),
            )
            return list(cursor.fetchall())


def _ensure_user_import_job_exists(job_id: str) -> None:
    if fetch_one(
        """
        SELECT id
        FROM user_import_job
        WHERE id = %s
        """,
        (job_id,),
    ) is None:
        raise ValueError("User import job not found")


def _serialize_user_import_row(row: dict) -> dict:
    normalized_values = dict(row["normalized_values"] or {})
    normalized_values.pop("password_hash", None)
    return {
        "id": str(row["id"]),
        "row_number": row["row_number"],
        "action": row["action"],
        "status": row["status"],
        "values": dict(row["values"] or {}),
        "normalized_values": normalized_values,
        "issues": list(row["issues"] or []),
        "existing_user": {
            "id": str(row["existing_user_id"]),
            "username": row["existing_user_username"],
            "email": row["existing_user_email"],
            "display_name": row["existing_user_display_name"],
            "status": row["existing_user_status"],
        }
        if row.get("existing_user_id") is not None
        else None,
    }


def _sanitize_user_import_row_for_response(row: dict) -> dict:
    normalized_values = dict(row["normalized_values"])
    normalized_values.pop("password_hash", None)
    return {**row, "normalized_values": normalized_values}


def _serialize_existing_user(existing_user: dict | None) -> dict | None:
    if existing_user is None:
        return None
    return {
        "id": str(existing_user["id"]),
        "username": existing_user["username"],
        "email": existing_user.get("email"),
        "display_name": existing_user["display_name"],
        "status": existing_user["status"],
        "role_codes": list(existing_user.get("role_codes") or []),
        "role_names": list(existing_user.get("role_names") or []),
    }


def _normalize_system_roles(system_roles: list[dict]) -> list[dict]:
    normalized = []
    for role in system_roles:
        if role.get("scope_kind") != "system":
            continue
        normalized.append(
            {
                "id": str(role["id"]),
                "code": _username_key(role.get("code")),
                "name": _trim_text(role.get("name")),
                "permissions": list(role.get("permissions") or []),
            }
        )
    normalized.sort(key=lambda item: item["code"])
    return normalized


def _normalize_filter(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip().lower()
    if normalized in {*ROW_STATUS_VALUES, *ROW_ACTION_VALUES}:
        return normalized
    raise ValueError("Unsupported import filter")


def _normalize_header(value: Any) -> str:
    normalized = _trim_text(value).lower()
    return LOCALIZED_IMPORT_HEADERS.get(_trim_text(value), normalized)


def _trim_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _username_key(value: Any) -> str:
    return _trim_text(value).lower()


def _normalize_email(value: Any) -> str | None:
    text = _trim_text(value).lower()
    return text or None


def _normalize_status(value: Any) -> str:
    text = _trim_text(value).lower()
    return text or "active"


def _normalize_password(value: Any) -> str:
    return _trim_text(value)


def _normalize_role_codes(value: Any) -> list[str]:
    if value is None:
        return []
    raw_values = value if isinstance(value, list) else re.split(r"[,;，；\n]+", str(value))
    normalized: list[str] = []
    seen: set[str] = set()
    for item in raw_values:
        code = _username_key(item)
        if code and code not in seen:
            seen.add(code)
            normalized.append(code)
    return normalized


def _password_hash_or_none(normalized_values: dict | None) -> str | None:
    if not normalized_values:
        return None
    password_hash = normalized_values.get("password_hash")
    return _trim_text(password_hash) or None


def _row_is_empty(values: dict[str, Any]) -> bool:
    return not any(_trim_text(value) for value in values.values())


def _issue(code: str, field: str, message: str, *, severity: Literal["error", "warning"] = "error") -> dict:
    return {"code": code, "field": field, "message": message, "severity": severity}


def _string_value(value: Any) -> str | None:
    if value is None:
        return None
    return str(value)
